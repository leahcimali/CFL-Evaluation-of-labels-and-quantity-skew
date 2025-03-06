import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Border, Side, Font

def load_and_prepare_data(skew):
    """
    Load and prepare the data based on the skew type.
    """
    if skew == 'quantity-skew-type-1':
        df = pd.read_csv('granular_results/qs1.csv')
        df.rename(columns={'global_accuracy': 'accuracy', 'homogeneity': 'hom', 'completeness': 'cmplt', 'v_measure': 'vm'}, inplace=True)
        granular_metrics = ['class_1_qt-skew_0.05', 'class_2_qt-skew_0.05', 'class_3_qt-skew_0.05', 'class_4_qt-skew_0.05',
                            'class_1_qt-skew_0.2', 'class_2_qt-skew_0.2', 'class_3_qt-skew_0.2', 'class_4_qt-skew_0.2',
                            'class_1_qt-skew_1', 'class_2_qt-skew_1', 'class_3_qt-skew_1', 'class_4_qt-skew_1',
                            'class_1_qt-skew_2', 'class_2_qt-skew_2', 'class_3_qt-skew_2', 'class_4_qt-skew_2']
    elif skew == 'quantity-skew-type-2':
        df = pd.read_csv('granular_results/qs2.csv')
        df.rename(columns={'global_accuracy': 'accuracy', 'homogeneity': 'hom', 'completeness': 'cmplt', 'v_measure': 'vm'}, inplace=True)
        granular_metrics = ['class_1_qt-skew_0.05', 'class_4_qt-skew_0.2', 'class_2_qt-skew_1', 'class_3_qt-skew_2']
    else:
        df = pd.read_csv('results/summarized_results.csv')
        granular_metrics = []

    for metric in granular_metrics:
        df[metric] = df[metric].str.split(' ').str[0].astype(float)

    return df, granular_metrics

def filter_and_clean_data(df, dataset, skew, heterogeneity_class):
    """
    Filter and clean the data based on the input parameters.
    """
    metrics = ['accuracy', 'ARI', 'AMI', 'hom', 'cmplt', 'vm']
    df.fillna('None', inplace=True)
    df = df[df['exp_type'] != 'oracle-centralized']
    df['exp_type'] = df['exp_type'] + '_' + df['params']

    if dataset:
        df = df[df['dataset'] == dataset]
    if skew:
        if skew == 'None':
            df = df[df['skew'] == skew]
        elif skew == 'quantity-skew':
            df = df[df['skew'] != 'None']
        else:
            df = df[df['skew'] == skew]
    if heterogeneity_class:
        df = df[df['heterogeneity_class'] == heterogeneity_class]

    df['accuracy'] = df['accuracy'].str.split(' ').str[0].astype(float)
    df[metrics] = df[metrics].apply(pd.to_numeric, errors='coerce')
    df['group'] = df['dataset'] + '_' + df['heterogeneity_class'] + '_' + df['skew'] + '_' + df['seed'].astype(str)
    df = df.drop(columns=['dataset', 'heterogeneity_class', 'skew', 'seed'])

    return df, metrics

def rank_algorithms(group, metrics):
    """
    Rank algorithms within each group for each metric.
    """
    for metric in metrics:
        group[f'{metric}_rank'] = group[metric].rank(ascending=False, method='min')
    return group

def calculate_best(df, metrics, granular_metrics):
    """
    Calculate the best algorithms based on the ranking.
    """
    grouped = df.groupby('group')
    ranked_groups = grouped.apply(rank_algorithms, metrics=metrics)
    best = ranked_groups.groupby('exp_type')[[f'{metric}_rank' for metric in metrics]].mean()
    best = best.sort_values(by='accuracy_rank')

    avg_accuracy = ranked_groups.groupby('exp_type')['accuracy'].mean()
    std_accuracy = ranked_groups.groupby('exp_type')['accuracy'].std()
    best['avg_accuracy'] = avg_accuracy
    best['std_accuracy'] = std_accuracy
    avg_ARI = ranked_groups.groupby('exp_type')['ARI'].mean()
    best['avg_ARI'] = avg_ARI

    if granular_metrics:
        for metric in granular_metrics:
            avg_metric = ranked_groups.groupby('exp_type')[metric].mean()
            best[f'{metric}'] = avg_metric

    best = best[['accuracy_rank', 'avg_accuracy', 'std_accuracy', 'ARI_rank', 'avg_ARI', 'AMI_rank', 'hom_rank', 'cmplt_rank', 'vm_rank']]
    return best

def algo_ranking(dataset=None, skew=None, heterogeneity_class=None):
    """
    Main function to perform algorithm ranking.
    """
    df, granular_metrics = load_and_prepare_data(skew)
    df, metrics = filter_and_clean_data(df, dataset, skew, heterogeneity_class)
    best = calculate_best(df, metrics, granular_metrics)
    return best

def write_dataframe_to_sheet(ws, df, start_row):
    """
    Write a dataframe to the worksheet with formatting.
    """
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    best_accuracy = df['avg_accuracy'].max()
    second_best_accuracy = df['avg_accuracy'][df['avg_accuracy'] != best_accuracy].max()

    for r_idx, row in enumerate(dataframe_to_rows(df.round(2), index=True, header=True), start_row):
        for c_idx, value in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            cell.border = thin_border
            if r_idx == start_row + 1:
                cell.font = Font(bold=True)
            if c_idx == df.columns.get_loc('avg_accuracy') + 2:
                if value == best_accuracy:
                    cell.font = Font(bold=True)
                elif value == second_best_accuracy:
                    cell.font = Font(underline='single')

    return r_idx + 2

def main():
    """
    Main function to execute the ranking and write results to Excel.
    """
    result = algo_ranking()
    results_qt1 = algo_ranking(skew='quantity-skew-type-1')
    results_qst2 = algo_ranking(skew='quantity-skew-type-2')
    results_noqs = algo_ranking(skew='None')

    wb = load_workbook('Results.xlsx')
    ws = wb.create_sheet(title='Global_results')

    ws.append(['Global_ranking'])
    ws.cell(row=ws.max_row, column=1).font = Font(bold=True)
    start_row = write_dataframe_to_sheet(ws, result, ws.max_row + 1)

    ws.append(['No Quantity Skew'])
    ws.cell(row=ws.max_row, column=1).font = Font(bold=True)
    write_dataframe_to_sheet(ws, results_noqs, ws.max_row + 1)

    ws.append(['Quantity skew type 1'])
    ws.cell(row=ws.max_row, column=1).font = Font(bold=True)
    start_row = write_dataframe_to_sheet(ws, results_qt1, ws.max_row + 1)

    ws.append(['Quantity skew type 2'])
    ws.cell(row=ws.max_row, column=1).font = Font(bold=True)
    start_row = write_dataframe_to_sheet(ws, results_qst2, ws.max_row + 1)

    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

    wb.save('Results.xlsx')

if __name__ == "__main__":
    main()
