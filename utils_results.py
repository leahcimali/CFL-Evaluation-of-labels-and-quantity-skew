from pandas import DataFrame
from pathlib import Path
from torch import tensor


def save_histograms(base_path: str) -> None:

    """Read csv files found in the given base_path and generates and saves histogram plots of clients assignments

    Arguments:
        base_path : The base directory path where the csv files are located

    Raises : 
        Warning when the csv file is not of the expected format (code generated results csv)
    """

    import pandas as pd

    pathlist = Path(base_path).rglob('*.csv') 
    
    for file_path in pathlist:

        if 'benchmark' not in str(file_path):
            
            try:

                df_results = pd.read_csv(file_path)

                plot_histogram_clusters(df_results, file_path.stem, base_path)
    
            except Exception as e:
        
                print(f"Error: Unable to open result file {file_path}.", e)
            
                continue
    return



def get_clusters(df_results : DataFrame) -> list:
    
    """ Function to returns a list of clusters ranging from 0 to max_cluster (uses: append_empty_clusters())
    """

    list_clusters = list(df_results['cluster_id'].unique())

    list_clusters = append_empty_clusters(list_clusters)

    return list_clusters


def append_empty_clusters(list_clusters : list) -> list:
    """
    Utility function for ``get_clusters'' to handle the situation where some clusters are empty by appending the clusters ID
    
    Arguments:
        list_clusters: List of clusters with clients

    Returns:
        List of clusters with or without clients
    """

    list_clusters_int = [int(x) for x in list_clusters]
    
    max_clusters = max(list_clusters_int)
    
    for i in range(max_clusters + 1):
        
        if i not in list_clusters_int:
            
            list_clusters.append(str(i))

    return list_clusters



def get_z_nclients(df_results : dict, x_het : list, y_clust : list, labels_heterogeneity : list) -> list:
    
    """ Returns the number of clients associated with a given heterogeneity class for each cluster"""

    z_nclients = [0]* len(x_het)

    for i in range(len(z_nclients)):
        
        z_nclients[i] = len(df_results[(df_results['cluster_id'] == y_clust[i]) &
                                       (df_results['heterogeneity_class'] == labels_heterogeneity[x_het[i]])])

    return z_nclients



def plot_img(img : tensor) -> None:

    """Utility function to plot an image of any shape"""

    from torchvision import transforms
    import matplotlib.pyplot as plt

    plt.imshow(transforms.ToPILImage()(img))



def plot_histogram_clusters(df_results: DataFrame, title: str, base_path: str) -> None:
    """ Function to create 3D Histograms of clients to cluster assignments showing client's heterogeneity class
    with skew values represented as hue in a stacked bar format.
    
    Arguments:
        df_results : DataFrame containing all parameters from the resulting csv files
        title : The plot title. The image is saved in base_path/plots/histogram_' + title + '.png'
        base_path : The base directory path where the plots will be saved
    """
    
    import matplotlib.pyplot as plt
    from matplotlib.patches import Patch
    import numpy as np

    # Unique values for heterogeneity_class and skew
    labels_heterogeneity = list(df_results['heterogeneity_class'].unique())
    skew_values = list(df_results['skew'].unique())
    
    bar_width = bar_depth = 0.3

    n_clusters = len(get_clusters(df_results))
    n_heterogeneities = len(labels_heterogeneity)

    # Create a color map based on skew values
    cmap = plt.get_cmap('tab10')  # Choosing a color map
    color_map = {skew: cmap(i / len(skew_values)) for i, skew in enumerate(skew_values)}

    # Set up the figure and axis
    fig = plt.figure()
    ax = fig.add_subplot(111, projection='3d')

    # Iterate over heterogeneity classes and clusters to create bars
    for h_idx, heterogeneity_class in enumerate(labels_heterogeneity):
        for cluster_id in get_clusters(df_results):
            
            # Filter data for the specific heterogeneity_class and cluster_id
            filtered_data = df_results[
                (df_results['heterogeneity_class'] == heterogeneity_class) & 
                (df_results['cluster_id'] == cluster_id)
            ]
            
            # If there's no data for this combination, skip
            if filtered_data.empty:
                continue

            # Initialize the bottom of the stack
            z_bottom = 0
            
            # Iterate over each skew value
            for skew in skew_values:
                # Count the number of clients for this skew value
                count = len(filtered_data[filtered_data['skew'] == skew])
                
                if count > 0:
                    # Coordinates for the current segment
                    x = h_idx
                    y = int(cluster_id)
                    dx = bar_width
                    dy = bar_depth
                    dz = count  # Height of the current segment
                    
                    # Draw the segment
                    ax.bar3d(x, y, z_bottom, dx, dy, dz, color=color_map[skew])
                    ax.view_init(elev=50)  # Change elev and azim for rotation
                    # Update the bottom for the next segment
                    z_bottom += dz

    # Setting up ticks and labels
    list_clusters = [x for x in 'abcdefghijklmnopqrstuvwxyz'][:n_clusters]

    ticksy = np.arange(0.25, len(list_clusters), 1)
    ticksx = np.arange(0.25, len(labels_heterogeneity), 1)

    plt.xticks(ticksx, labels_heterogeneity)
    plt.yticks(ticksy, list_clusters)

    plt.ylabel('Cluster ID')
    plt.xlabel('Heterogeneity Class')
    
    ax.set_zlabel('Number of Clients')
    
    plt.title(title, fontdict=None, loc='center', pad=None)
    
    # Create a legend for the color mapping
    legend_patches = [Patch(color=color_map[skew], label=skew) for skew in skew_values]
    plt.legend(handles=legend_patches, title="Skew", loc='upper right', bbox_to_anchor=(1.15, 1))
    
    plt.savefig(Path(base_path) / 'plots' / f'histogram_{title}.png')
    
    plt.close()

    return




def normalize_results(results_accuracy : float, results_std : float) -> int:

    """Utility function to convert float accuracy and std to percentage """
    
    if results_accuracy < 1:
        
        results_accuracy = results_accuracy * 100
    
        results_std = results_std * 100

    return results_accuracy, results_std


def summarize_results(base_path) -> None:

    """ Creates results summary of all the results files under "results/summarized_results.csv"""
        
    from pathlib import Path
    import pandas as pd
    from numpy import mean, std

    from src.metrics import calc_global_metrics
    import sys



    pathlist = Path(base_path).rglob('*.csv')
    list_results = []

    for path in pathlist:
        
        if 'summarized_results' not in str(path):
            
            df_exp_results = pd.read_csv(path)

            results_accuracy = mean(list(df_exp_results['accuracy'])) 
            results_std = std(list(df_exp_results['accuracy']))

            results_accuracy, results_std = normalize_results(results_accuracy, results_std)
            num_clusters = len(df_exp_results['cluster_id'].unique())
            accuracy =  "{:.2f}".format(results_accuracy) + " \\pm " +   "{:.2f}".format(results_std)

            list_params = path.stem.split('_')      

            dict_exp_results = {
            "exp_type": list_params[0],
            "params": list_params[1],
            "dataset": list_params[2],
            "seed" : list_params[11],
            "nn_model": list_params[3],
            "heterogeneity_class": list_params[4],
            "skew": list_params[5],
            "number_of_clients": list_params[6],
            "samples by_client": list_params[7],
            "num_clusters": num_clusters,
            "epochs": list_params[9],
            "rounds": list_params[10],
            "accuracy": accuracy
            }

            try:
                
                labels_true = list(df_exp_results['heterogeneity_class'])
                labels_pred = list(df_exp_results['cluster_id'])
                
                dict_metrics = calc_global_metrics(labels_true=labels_true, labels_pred=labels_pred)
            
                dict_exp_results.update(dict_metrics)
                
            
            except:

                print(f"Warning: Could not calculate cluster metrics for file {path}")
                
    
            list_results.append(dict_exp_results)
            
    df_results = pd.DataFrame(list_results)
    df_results.sort_values(['heterogeneity_class', 'dataset', 'seed','exp_type', 'nn_model','number_of_clients'], inplace=True)

    try: 
        df_results = df_results[['exp_type', "params", 'nn_model', 'number_of_clients', 'dataset', 'num_clusters', 'heterogeneity_class', 'skew', "accuracy", "ARI", "AMI", "hom", "cmplt", "vm","seed"]]
    except KeyError as e: 
        missing_cols = [col for col in ["ARI", "AMI", "hom", "cmplt", "vm"] if col not in df_results.columns]
        for col in missing_cols:
            df_results[col] = "n/a"
        df_results = df_results[['exp_type', "params", 'nn_model', 'number_of_clients', 'dataset', 'num_clusters', 'heterogeneity_class', 'skew', "accuracy", "ARI", "AMI", "hom", "cmplt", "vm","seed"]]
    df_results.sort_values(['dataset', 'skew','exp_type','accuracy'], inplace=True)

    df_results.to_csv(Path(base_path) / "summarized_results.csv", float_format='%.2f', index=False, na_rep="n/a")

    return


def granular_results(base_path) -> None : 
    """
    Processes experimental results to compute and save accuracy statistics by class and skew type.
    Parameters:
    df_exp_results (pd.DataFrame): DataFrame containing experimental results with columns 'heterogeneity_class', 'skew', and 'accuracy'.
    dict_exp_results (dict): Dictionary containing experiment metadata with keys 'exp_type', 'dataset', 'heterogeneity_class', and 'skew'.
    Returns:
    None: The function saves the processed results to a CSV file based on the skew type.
    """
    import pandas as pd
    from pathlib import Path
    from numpy import mean, std
    from src.metrics import calc_global_metrics

    pathlist = [path for path in Path(base_path).rglob('*.csv') if 'summarized_results' not in str(path) and 'noqs.csv' not in str(path) and 'qs1.csv' not in str(path) and 'qs2.csv' not in str(path)]
    
    for path in pathlist:
        
        df_exp_results = pd.read_csv(path)

        results_accuracy = mean(list(df_exp_results['accuracy'])) 
        results_std = std(list(df_exp_results['accuracy']))
        num_clusters = len(df_exp_results['cluster_id'].unique())

        results_accuracy, results_std = normalize_results(results_accuracy, results_std)

        accuracy =  "{:.2f}".format(results_accuracy) + " \\pm " +   "{:.2f}".format(results_std)

        list_params = path.stem.split('_')      

        dict_exp_results = {
        "exp_type": list_params[0],
        "params": list_params[1],
        "dataset": list_params[2], 
        "seed": list_params[11],
        "nn_model": list_params[3],
        "heterogeneity_class": list_params[4],
        "skew": list_params[5],
        "number_of_clients": list_params[6],
        "samples by_client": list_params[7],
        "num_clusters": num_clusters,
        "epochs": list_params[9],
        "rounds": list_params[10],
        "accuracy": accuracy
        }

        try:
                
            labels_true = list(df_exp_results['heterogeneity_class'])
            labels_pred = list(df_exp_results['cluster_id'])
            
            dict_metrics = calc_global_metrics(labels_true=labels_true, labels_pred=labels_pred)
        
            
        
        except:

            print(f"Warning: Could not calculate cluster metrics for file {path}")
            
        accuracy_by_class = df_exp_results.groupby(['heterogeneity_class', 'skew'])['accuracy'].agg(['mean', 'std']).reset_index()
        # Rename columns for clarity
        accuracy_by_class.columns = ['heterogeneity_class', 'skew', 'mean_accuracy', 'std_accuracy']
        # Map heterogeneity_class to class_1, class_2, etc.
        class_mapping = {value: f'class_{i+1}' for i, value in enumerate(accuracy_by_class['heterogeneity_class'].unique())}
        accuracy_by_class['heterogeneity_class'] = accuracy_by_class['heterogeneity_class'].map(class_mapping)
        # Concatenate the heterogeneity_class and skew columns
        accuracy_by_class['class_skew'] = accuracy_by_class['heterogeneity_class'].astype(str) + '_' + accuracy_by_class['skew']
        accuracy_by_class['accuracy'] = accuracy_by_class['mean_accuracy'].round(2).astype(str) + ' \pm ' + accuracy_by_class['std_accuracy'].round(2).astype(str)
        accuracy_by_class.drop(columns=['heterogeneity_class', 'skew','mean_accuracy','std_accuracy'], inplace=True)
        df = accuracy_by_class.pivot_table(index=None, columns='class_skew', values='accuracy', aggfunc='first')
        df.columns.name = None
        df.reset_index(drop=True, inplace=True)
        # Add the exp_type, dataset, heterogeneity_class, and skew to the dataframe
        df['exp_type'] = dict_exp_results['exp_type']
        df['params'] = dict_exp_results['params']
        df['dataset'] = dict_exp_results['dataset']
        df['heterogeneity_class'] = dict_exp_results['heterogeneity_class']
        df['skew'] = dict_exp_results['skew']
        df['global_accuracy'] = accuracy
        df['seed'] = dict_exp_results['seed']
        
        if   df['exp_type'].isin(['fedavg', 'fedprox']).any():     
            df['ARI'] = 0.0
            df['AMI'] = 0.0        
            df['homogeneity'] = 0.0
            df['completeness'] = 0.0
            df['v_measure'] = 0.0
        elif df['exp_type'].isin(['oracle-centralized']).any():
            df['ARI'] = 1.0
            df['AMI'] = 1.0        
            df['homogeneity'] = 1.0
            df['completeness'] = 1.0
            df['v_measure'] = 1.0
        else : 
            df['ARI'] = dict_metrics['ARI']
            df['AMI'] = dict_metrics['AMI']
            df['homogeneity'] = dict_metrics['hom']
            df['completeness'] = dict_metrics['cmplt']
            df['v_measure'] = dict_metrics['vm']
         

        # Reorder columns to place the new columns at the beginning
        cols = ['dataset', 'heterogeneity_class','seed','exp_type','params','skew'] + [col for col in df.columns if col not in ['exp_type','params', 'dataset', 'heterogeneity_class','seed', 'skew']]
        df = df[cols]
        granular_results_path = Path('granular_results')
        granular_results_path.mkdir(parents=True, exist_ok=True)
        if dict_exp_results['skew'] == 'quantity-skew-type-1':
            file_path = "granular_results/qs1.csv"
        elif dict_exp_results['skew'] == 'quantity-skew-type-2':
            file_path = "granular_results/qs2.csv"
        else:
            file_path = "granular_results/noqs.csv"

        if Path(file_path).exists():
            df.to_csv(file_path, mode='a', header=False, float_format='%.2f', index=False, na_rep="n/a")
        else:
            df.to_csv(file_path, float_format='%.2f', index=False, na_rep="n/a")
    return

import pandas as pd
import glob
import os
from openpyxl.styles import Font, Border, Side, Color
from openpyxl.utils import get_column_letter

# Function to extract the first number from the 'global_accuracy' string
def extract_number1(global_accuracy):
    if '\\pm' in global_accuracy:
        return float(global_accuracy.split('\\pm')[0].strip())
    else:
        return float(global_accuracy.strip())

# Function to auto-adjust column widths
def auto_adjust_column_widths(sheet, df):
    for col in df.columns:
        max_length = max(len(str(col)), df[col].astype(str).map(len).max())
        sheet.column_dimensions[get_column_letter(df.columns.get_loc(col) + 1)].width = max_length + 2

# Function to apply formatting to the highest and second-highest global accuracy values
def format_global_accuracy(sheet, df_sorted, group):
    global_accuracies = group['global_accuracy'].apply(extract_number1).tolist()
    indices = group.index.tolist()
    
    sorted_indices = sorted(indices, key=lambda x: extract_number1(df_sorted.loc[x, 'global_accuracy']), reverse=True)
    highest_index = sorted_indices[0]
    second_highest_index = sorted_indices[1] if len(sorted_indices) > 1 else None
    
    excel_row_highest = df_sorted.index.get_loc(highest_index) + 2
    excel_row_second_highest = df_sorted.index.get_loc(second_highest_index) + 2 if second_highest_index is not None else None
    
    cell = sheet.cell(row=excel_row_highest, column=df_sorted.columns.get_loc('global_accuracy') + 1)
    cell.font = Font(bold=True)
    
    if second_highest_index is not None:
        cell = sheet.cell(row=excel_row_second_highest, column=df_sorted.columns.get_loc('global_accuracy') + 1)
        cell.font = Font(bold=True, underline='single', color='585858')

# Function to apply formatting to ARI values
def format_ari_values(sheet, df_sorted, group):
    filtered_group = group[group['ARI'] < 1]
    
    if not filtered_group.empty:
        sorted_indices = filtered_group.sort_values(by='ARI', ascending=False).index.tolist()
        highest_index_ARI = sorted_indices[0]
        excel_row_highest_ARI = df_sorted.index.get_loc(highest_index_ARI) + 2
        cell = sheet.cell(row=excel_row_highest_ARI, column=df_sorted.columns.get_loc('ARI') + 1)
        cell.font = Font(bold=True, underline='single', color='585858')

# Function to apply borders and formatting to specific columns
def apply_borders_and_formatting(sheet, df_sorted, group):
    thick_black_border = Border(bottom=Side(style='thick', color='000000'))
    thick_vertical_border = Border(right=Side(style='thick', color='000000'))
    
    last_row_index = group.index[-1]
    excel_last_row = df_sorted.index.get_loc(last_row_index) + 2
    for col_name in df_sorted.columns:
        col_idx = df_sorted.columns.get_loc(col_name) + 1
        sheet.cell(row=excel_last_row, column=col_idx).border = thick_black_border

    for row_idx, row in df_sorted.iterrows():
        excel_row = df_sorted.index.get_loc(row_idx) + 2
        ari_value = row['ARI']
        if ari_value == 1:
            col_idx = df_sorted.columns.get_loc('ARI') + 1
            cell = sheet.cell(row=excel_row, column=col_idx)
            cell.font = Font(bold=True)
        for col_name in ['seed', 'params', 'ARI']:
            col_idx = df_sorted.columns.get_loc(col_name) + 1
            sheet.cell(row=excel_row, column=col_idx).border = thick_vertical_border

# Function to apply formatting to cells with values below 50
def format_low_values(sheet, df_sorted):
    exclude_columns = {'AMI', 'homogeneity', 'completeness', 'v_measure'}
    
    for row_idx, row in df_sorted.iterrows():
        excel_row = df_sorted.index.get_loc(row_idx) + 2  # Excel rows start from 1, and header is row 1
        ari_col_idx = df_sorted.columns.get_loc('ARI') + 1  # Column index of 'ARI' in Excel
        
        # Get the value of column D (Excel column index 4) for the current row
        exp_type_value = sheet.cell(row=excel_row, column=4).value
        
        # Check if the value of column D is not in the excluded list
        if exp_type_value not in ['\nfedavg', 'fedavg','fedprox', 'oracle-centralized']:
            for col_idx in range(ari_col_idx + 1, len(df_sorted.columns) + 1):
                col_name = df_sorted.columns[col_idx - 1]
                if col_name in exclude_columns:
                    continue
                
                cell_value = row[col_name]
                try:
                    extracted_value = extract_number1(str(cell_value))
                except (ValueError, AttributeError):
                    continue
                
                # Apply formatting based on extracted_value
                if extracted_value < 50:
                    cell = sheet.cell(row=excel_row, column=col_idx)
                    cell.font = Font(bold=True, color="FF0000")
                elif extracted_value < 70:
                    cell = sheet.cell(row=excel_row, column=col_idx)
                    cell.font = Font(bold=True, color="800080")

# Main program
def main_excel():
    with pd.ExcelWriter('granular_results.xlsx', engine='openpyxl') as writer:
        for csv_file in glob.glob('granular_results/*.csv'):
            df = pd.read_csv(csv_file)
            df = df.drop(columns=['skew'])

            if os.path.basename(csv_file) == 'noqs.csv':
                sheet_name = 'noqs'
                df = df[['dataset', 'heterogeneity_class','seed', 'exp_type', 'params', 'global_accuracy', 'ARI',
                         'class_1_none', 'class_2_none', 'class_3_none', 'class_4_none', 'AMI', 'homogeneity', 
                         'completeness', 'v_measure']]
            elif os.path.basename(csv_file) == 'qs1.csv':
                sheet_name = 'qs1'
                df = df[['dataset', 'heterogeneity_class', 'seed','exp_type', 'params', 'global_accuracy', 'ARI', 
                         'class_1_qt-skew_0.05', 'class_2_qt-skew_0.05', 'class_3_qt-skew_0.05', 'class_4_qt-skew_0.05',
                         'class_1_qt-skew_0.2', 'class_2_qt-skew_0.2', 'class_3_qt-skew_0.2', 'class_4_qt-skew_0.2',
                         'class_1_qt-skew_1', 'class_2_qt-skew_1', 'class_3_qt-skew_1', 'class_4_qt-skew_1',
                         'class_1_qt-skew_2', 'class_2_qt-skew_2', 'class_3_qt-skew_2', 'class_4_qt-skew_2',
                         'AMI', 'homogeneity', 'completeness', 'v_measure']]
            elif os.path.basename(csv_file) == 'qs2.csv':
                sheet_name = 'qs2'
                df = df[['dataset', 'heterogeneity_class','seed', 'exp_type', 'params', 'global_accuracy', 'ARI', 
                        'class_1_qt-skew_0.05', 'class_4_qt-skew_0.2', 'class_2_qt-skew_1', 'class_3_qt-skew_2',
                        'AMI', 'homogeneity', 'completeness', 'v_measure']]
                
            exp_type_order = ['fedavg', 'fedprox', 'cfl', 'hcfl', 'fedgroup', 'ifca', 'srfca', 'cornflqs', 'oracle-centralized']
            df['exp_type'] = pd.Categorical(df['exp_type'], categories=exp_type_order, ordered=True)
            df['exp_type'] = df['exp_type'].apply(lambda x: '\n' + x if x == 'fedavg' else x)
            if df['exp_type'].isin(['\nfedavg','fedavg', 'fedprox']).any():
                df.loc[df['exp_type'].isin(['fedavg', 'fedprox']), ['ARI', 'AMI', 'homogeneity', 'completeness', 'v_measure']] = 0
            elif df['exp_type'].isin(['oracle-centralized']).any():
                df.loc[df['exp_type'].isin(['oracle-centralized']), ['ARI', 'AMI', 'homogeneity', 'completeness', 'v_measure']] = 1
                
            df_sorted = df.sort_values(['dataset', 'heterogeneity_class', 'seed','exp_type'])

            df_sorted.to_excel(writer, sheet_name=sheet_name, index=False)
            
            workbook = writer.book
            sheet = workbook[sheet_name]
            auto_adjust_column_widths(sheet, df_sorted)
            
            for _, group in df_sorted.groupby(['dataset', 'heterogeneity_class','seed']):
                format_global_accuracy(sheet, df_sorted, group)
                format_ari_values(sheet, df_sorted, group)
                apply_borders_and_formatting(sheet, df_sorted, group)
            
            format_low_values(sheet, df_sorted)
    
    print("Excel file created successfully!")

if __name__ == "__main__":
    import sys
    if len(sys.argv) > 1:
        base_path = sys.argv[1]
    else:
        base_path = "results/"

    save_histograms(base_path)
    summarize_results(base_path)
    granular_results(base_path)
    try : 
        main_excel()
    except :
        print('Excel generation not working on this system')